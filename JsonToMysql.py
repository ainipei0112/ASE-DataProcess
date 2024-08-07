import re
import os
import csv
import json
import time
import shutil
import chardet
import datetime
import numpy as np
import pandas as pd
import mysql.connector
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

# ----------------------------------- 函數定義 -----------------------------------

# Json 資料讀取設定
def read_Jsonfile(settings_path):
    try:
        print("Reading database.")
        with open(settings_path, "r", encoding='utf-8') as r_file:
            databases = json.load(r_file)
        database = databases["folder_details"]
        return database
    except Exception as e:
        print(f"Failed to read database: {e}")
        exit()

#設定Excel標題格式
def reset_ws():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '工作表1'

    # 標題欄位
    column_titles = ['Date', 'Date_1', 'Lot', 'AOI_ID', 'AOI_Scan_Amount', 'AOI_Pass_Amount', 'AOI_Reject_Amount', 'AOI_Yield', 'AOI_Yield_Die_Corner', 'AI_Pass_Amount', 'AI_Reject_Amount', 'AI_Yield', 'AI_Fail_Corner_Yield', 'Final_Pass_Amount', 'Final_Reject_Amount', 'Final_Yield', 'AI_EA_Overkill_Die_Corner', 'AI_EA_Overkill_Die_Surface', 'AI_Image_Overkill_Die_Corner', 'AI_Image_Overkill_Die_Surface', 'EA_over_kill_Die_Corner', 'EA_over_kill_Die_Surface', 'Image_Overkill_Die_Corner', 'Image_Overkill_Die_Surface', 'Total_Images', 'Image_Overkill', 'AI_Fail_EA_Die_Corner', 'AI_Fail_EA_Die_Surface', 'AI_Fail_Image_Die_Corner', 'AI_Fail_Image_Die_Surface', 'AI_Fail_Total', 'Total_AOI_Die_Corner_Image', 'AI_Pass', 'AI_Reduction_Die_Corner', 'AI_Reduction_All', 'True_Fail', 'True_Fail_Crack', 'True_Fail_Chipout', 'True_Fail_Die_Surface', 'True_Fail_Others', 'EA_True_Fail_Crack', 'EA_True_Fail_Chipout', 'EA_True_Fail_Die_Surface', 'EA_True_Fail_Others', 'EA_True_Fail_Crack_Chipout', 'Device_ID', 'OP_EA_Die_Corner', 'OP_EA_Die_Surface', 'OP_EA_Others', 'Die_Overkill']

    for i, title in enumerate(column_titles, start=1):
        cell = ws1.cell(row=1, column=i, value=title)
        ws1.column_dimensions[get_column_letter(i)].width = 22
        fill_color = '95B3D7' if i < 21 else ('FDE9D9' if i < 27 else 'B7DEE8')
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        cell.fill = fill
        side1 = Side(color='000000', style='thin')
        cell.font = Font(name='Microsoft YaHei', size=12)
        cell.alignment = Alignment(vertical='center', horizontal='center')
        cell.border = Border(left=side1, right=side1, top=side1, bottom=side1)

    ws1.row_dimensions[1].height = 49.5
    return wb, ws1

# 處理 JSON 資料並寫入 Excel
def JsonToExcel(database, main_path, output_path, yesterday, today, wb, ws1, weekend="", output_type="both"):
    # 初始化垃圾文件和Excel文件列表
    trash = []
    excel_files = []
    list_data = []
    list_data2 = []
    list_data_t = []

    # 日期格式轉換
    date_yesterday = datetime.datetime.strptime(str(datetime.datetime.now().year) + yesterday, "%Y%m%d")
    date_today = datetime.datetime.strptime(str(datetime.datetime.now().year) + today, "%Y%m%d")
    yesterday_str = datetime.datetime.strftime(date_yesterday, '%Y-%m-%d 07:30:00')  # 只抓取當天七點半後的資料
    today_str = datetime.datetime.strftime(date_today, '%Y-%m-%d 07:30:00')
    yesterday1 = datetime.datetime.strptime(yesterday_str, '%Y-%m-%d %H:%M:%S')
    today1 = datetime.datetime.strptime(today_str, '%Y-%m-%d %H:%M:%S')

    # 主路徑下所有的文件夾及文件
    directories = [f.path for f in os.scandir(main_path)]
    for date_directory in directories:
        if os.path.isfile(date_directory):
            filename = date_directory.split("\\")[-1]
            if date_directory.endswith("csv"):
                excel_files.append(filename.split("_")[-2])
            trash.append(date_directory)

    for date_directory in directories:
        if os.path.isdir(date_directory):
            directory_name = date_directory.split("\\")[-1]
            try:
                date_folder = datetime.datetime.strptime(directory_name, "%Y-%m-%d")
                if date_folder < date_yesterday or date_folder > date_today:
                    trash.append(date_directory)
                    continue
            except:
                pass

    # 清理垃圾文件
    directories = [d for d in directories if d not in trash]  # 使用列表推導式替換刪除迴圈

    # 將所有 Json 進行邏輯運算
    for directory in directories: # 所有日期的資料夾
        lot_names = [f.path for f in os.scandir(directory) if os.path.isdir(f.path)]

        for lot_name in lot_names: # 所有批號的資料夾
            print(lot_name)
            Json_files = [f.path for f in os.scandir(lot_name)if os.path.isfile(f.path)]

            for Json_file in Json_files:
                try:
                    with open(Json_file) as json_file:
                        data = json.load(json_file)
                except:
                    pass

                if "OP_Checked" in data : # 確認 JSON 是否 "OP_Checked"
                    if data["AOI_Scan_Amount"] != 0 : # 確認 AOI 掃描數量不為 0
                        incorrect_mag_counter = 0
                        chipout_counter = 0
                        crack_counter = 0
                        others_counter = 0
                        op_incorrect_mag_counter = 0
                        op_chipout_counter = 0
                        op_crack_counter = 0
                        op_others_counter  = 0
                        corner_duplicate_Checker = {}
                        die_duplicate_Checker = {}
                        duplicate_Checker = {}
                        xdie=[]

                        # 獲取 OP_PassDetails、OP_FailDetails、AI_FailDetails 中的 id
                        for file_data in data["OP_PassDetails"]:
                            xdie.append(file_data["id"])
                        for file_data in data["OP_FailDetails"]:
                            xdie.append(file_data["id"])
                        for file_data in data["AI_FailDetails"]:
                            if not file_data["id"] in xdie:
                                continue

                            # 根據不同的 aoidefecttype 進行計數
                            if(file_data["aoiDefectType"] == "Incorrect_Magnification" or file_data["aoiDefectType"] == "Incorrect_Size" or file_data["aoiDefectType"] =="Scratch" or file_data["aoiDefectType"] =="Passivation_Effect" or file_data["aoiDefectType"] =="OP_Ink"):
                                incorrect_mag_counter += 1

                                for tempdata in data["OP_PassDetails"]:
                                    if(file_data["id"] == tempdata["id"]):
                                        op_incorrect_mag_counter +=1
                                        break
                            elif(file_data["aoiDefectType"] == "chipout" or file_data["aoiDefectType"] == "Chipout" or file_data["aoiDefectType"] == "Peeling"):
                                chipout_counter +=1

                                for tempdata in data["OP_PassDetails"]:
                                    if(file_data["id"] == tempdata["id"]):
                                        op_chipout_counter +=1
                                        break
                            elif(str(file_data["aoiDefectType"]) == "Crack"):
                                crack_counter +=1

                                for tempdata in data["OP_PassDetails"]:
                                    if(file_data["id"] == tempdata["id"]):
                                        op_crack_counter +=1
                                        break
                            else:
                                others_counter +=1

                                for tempdata in data["OP_PassDetails"]:
                                    if(file_data["id"] == tempdata["id"]):
                                        op_others_counter +=1
                                        break

                            # 獲取 XY 座標
                            XY = file_data["fileName"].split("_")[5:9]
                            XY = "_".join(XY)

                            # 根據不同的 aoidefecttype 更新 die_duplicate_Checker 或 corner_duplicate_Checker
                            if(file_data["aoiDefectType"] == "Incorrect_Magnification" or file_data["aoiDefectType"] == "Incorrect_Size" or file_data["aoiDefectType"] =="Scratch" or file_data["aoiDefectType"] =="Passivation_Effect" or file_data["aoiDefectType"] =="OP_Ink"):
                                if(XY in die_duplicate_Checker):
                                    die_duplicate_Checker[XY] +=1
                                else:
                                    die_duplicate_Checker[XY] = 1
                            else:
                                if(XY in corner_duplicate_Checker):
                                    corner_duplicate_Checker[XY] +=1
                                else:
                                    corner_duplicate_Checker[XY] = 1

                        EA_Fail_die = len(die_duplicate_Checker)
                        EA_Fail_corner = len(corner_duplicate_Checker)
                        OP_ChipOut={}
                        OP_Metal_Scratch={}
                        OP_Others={}

                        for file_data in data["OP_FailDetails"]:
                            XY = file_data["fileName"].split("_")[5:9]
                            XY = "_".join(XY)
                            try:
                                if XY in die_duplicate_Checker:
                                    del die_duplicate_Checker[XY]
                                if XY in corner_duplicate_Checker:
                                    del corner_duplicate_Checker[XY]
                                if file_data["opRejudgeDefectMode"]=="ChipOut":
                                    if(XY in OP_ChipOut):
                                        OP_ChipOut[XY] +=1
                                    else:
                                        OP_ChipOut[XY] = 1
                                elif file_data["opRejudgeDefectMode"]=="Metal Scratch":
                                    if(XY in OP_Metal_Scratch):
                                        OP_Metal_Scratch[XY] +=1
                                    else:
                                        OP_Metal_Scratch[XY] = 1
                                else:
                                    if(XY in OP_Others):
                                        OP_Others[XY] +=1
                                    else:
                                        OP_Others[XY] = 1
                            except:
                                pass

                        print(Json_file)
                        OG_loc = Json_file.split("\\")
                        date_obj = datetime.datetime.strptime(OG_loc[6], "%Y-%m-%d")
                        date_str = date_obj.strftime("%m%d%Y")
                        OG_Loc='\\\\'+OG_loc[2]+'\\'+OG_loc[3]+'\\'+OG_loc[4]+'\\Image'+'\\'+date_str+'\\'+OG_loc[7]+'\\'+OG_loc[8].replace('.json','')

                        # 輸出成功顯示OK
                        if not os.path.exists(OG_Loc):
                            print('DIE',OG_Loc)
                            break
                        else:
                            print('OK',OG_Loc)

                        for file_data in os.scandir(OG_Loc):
                            if(file_data.name.endswith("jpg")):
                                XY = file_data.name.split("_")[5:9]
                                XY = "_".join(XY)
                                if(XY in duplicate_Checker):
                                    duplicate_Checker[XY] +=1
                                else:
                                    duplicate_Checker[XY] = 1

                        AOI_Fail = len(duplicate_Checker)
                        ai_total_image = (int(data["AI_Pass"]) + int(data["AI_Fail"]))
                        total_overkill = op_crack_counter + op_chipout_counter + op_incorrect_mag_counter + op_others_counter
                        total_ai_fail = incorrect_mag_counter + chipout_counter + crack_counter + others_counter

                        if(len(data["AI_PassDetails"]) == 0):
                            ai_reduction_percent = 1
                            ai_reduction_Sawpercent = 1
                        else:
                            ai_reduction_percent = float(len(data["AI_PassDetails"]))/float(ai_total_image)
                            ai_reduction_Sawpercent = float(data["AI_Pass"])/float(data["AI_Pass"] + crack_counter + chipout_counter + others_counter)

                        if(int(data["Reject_Amount"]) == 0):
                            AI_Yield = 1
                        else:
                            AI_Yield = 1-(float(data["Reject_Amount"]) / float(data["AOI_Scan_Amount"]))

                        if total_ai_fail-incorrect_mag_counter == 0:
                            Ai_Overkill_Saw = 0
                        else:
                            Ai_Overkill_Saw = float(abs(total_overkill - op_incorrect_mag_counter)) / float(ai_total_image)

                        if incorrect_mag_counter == 0:
                            incorrect_mag_overkill_per = 0
                        else:
                            incorrect_mag_overkill_per = float(op_incorrect_mag_counter)/float(ai_total_image)

                        data_filename = Json_file.split("\\")[-1]
                        data_filename = data_filename.split(".")[:-1]
                        data_filename = ".".join(data_filename)

                        print(data_filename)
                        bre = False
                        for database_data in database:
                            if not (data_filename in database_data["filename"]):
                                pass
                            else:
                                print("IN!")
                                bre = True
                                break

                        if not (bre):
                            continue

                        for database_data in database:
                            if(database_data["filename"] == data_filename):
                                data_date = database_data["Date"]
                                data_device_ID = database_data["Device_ID"]
                                break

                        EA_OP_Crack = 0
                        EA_OP_Chipout = 0
                        EA_OP_DieSurface = 0
                        EA_OP_Others = 0
                        EA_OP_Duplicate_Checker = {}
                        True_Fail_Crack = 0
                        True_Fail_Chipout = 0
                        True_Fail_Die_Surface = 0
                        True_Fail_Others = 0

                        for file_data in data["AI_FailDetails"]:
                            fail = False
                            XY = file_data["fileName"].split("_")[5:9]
                            XY = "_".join(XY)

                            for OP_file_data in data["OP_FailDetails"]:
                                OP_XY = OP_file_data["fileName"].split("_")[5:9]
                                OP_XY = "_".join(OP_XY)
                                if XY == OP_XY:
                                    fail = True
                                    break

                            if fail:
                                if XY in EA_OP_Duplicate_Checker:
                                    if file_data["aoiDefectType"] in ("Incorrect_Magnification", "Incorrect_Size", "Scratch", "Passivation_Effect", "OP_Ink"):
                                        True_Fail_Die_Surface += 1
                                    elif file_data["aoiDefectType"] in ("chipout", "Chipout", "Peeling"):
                                        True_Fail_Chipout += 1
                                    elif file_data["aoiDefectType"] == "Crack":
                                        True_Fail_Crack += 1
                                    else:
                                        True_Fail_Others += 1
                                else:
                                    EA_OP_Duplicate_Checker[XY] = 1

                                    if file_data["aoiDefectType"] in ("Incorrect_Magnification", "Incorrect_Size", "Scratch", "Passivation_Effect", "OP_Ink"):
                                        EA_OP_DieSurface += 1
                                        True_Fail_Die_Surface += 1
                                    elif file_data["aoiDefectType"] in ("chipout", "Chipout", "Peeling"):
                                        EA_OP_Chipout += 1
                                        True_Fail_Chipout += 1
                                    elif file_data["aoiDefectType"] == "Crack":
                                        True_Fail_Crack += 1
                                        EA_OP_Crack += 1
                                    else:
                                        EA_OP_Others += 1
                                        True_Fail_Others += 1

                        if int(data["AOI_Scan_Amount"]) == 0:
                            continue

                        date_time = datetime.datetime.strptime(data_date, "%Y-%m-%d %H:%M:%S").time()
                        compare_time = datetime.time(7, 30)

                        if date_time>=compare_time:
                            date_day=datetime.datetime.strptime(data_date, "%Y-%m-%d %H:%M:%S").date()
                        else:
                            date_day=(datetime.datetime.strptime(data_date, "%Y-%m-%d %H:%M:%S")-datetime.timedelta(1)).date()

                        data_dictionary = {
                            "Date" : data_date,
                            "Date_1" : date_day,
                            "Lot" : lot_name.split("\\")[-1],
                            "AOI_ID" : Json_file.split(".")[-2],
                            "AOI_Scan_Amount" : data["AOI_Scan_Amount"],
                            "AOI_Pass_Amount" : data["AOI_Scan_Amount"]-AOI_Fail,
                            "AOI_Reject_Amount" : AOI_Fail,
                            "AOI_Yield" :   float(data["AOI_Scan_Amount"]-AOI_Fail)/float(data["AOI_Scan_Amount"]),
                            "AOI_Yield_Die_Corner" :   float(data["AOI_Scan_Amount"]-(AOI_Fail-EA_Fail_die))/float(data["AOI_Scan_Amount"]),
                            "AI_Pass_Amount" : int(data["AOI_Scan_Amount"]) - (EA_Fail_corner + EA_Fail_die),
                            "AI_Reject_Amount" : EA_Fail_corner + EA_Fail_die,
                            "AI_Yield" :float(int(data["AOI_Scan_Amount"]) - (EA_Fail_corner + EA_Fail_die)) / float(data["AOI_Scan_Amount"]),
                            "AI_Fail_Corner_Yield" :float(data["AOI_Scan_Amount"] - EA_Fail_corner) / float(data["AOI_Scan_Amount"]),
                            "Final_Pass_Amount" : data["Pass_Amount"],
                            "Final_Reject_Amount" : data["Reject_Amount"],
                            "Final_Yield" : AI_Yield,
                            "AI_EA_Overkill_Die_Corner" : (float(len(corner_duplicate_Checker))/float(data["AOI_Scan_Amount"])) if len(corner_duplicate_Checker) != 0 else 0,
                            "AI_EA_Overkill_Die_Surface" : (float(len(die_duplicate_Checker))/float(data["AOI_Scan_Amount"]))if len(die_duplicate_Checker) != 0 else 0,
                            "AI_Image_Overkill_Die_Corner" : Ai_Overkill_Saw,
                            "AI_Image_Overkill_Die_Surface" : incorrect_mag_overkill_per,
                            "EA_over_kill_Die_Corner" : len(corner_duplicate_Checker),
                            "EA_over_kill_Die_Surface" : len(die_duplicate_Checker),
                            "Image_Overkill_Die_Corner" : abs(total_overkill - op_incorrect_mag_counter),
                            "Image_Overkill_Die_Surface" : op_incorrect_mag_counter,
                            "Total_Images" : ai_total_image,
                            "Image_Overkill" : abs(total_overkill - op_incorrect_mag_counter)+op_incorrect_mag_counter,
                            "AI_Fail_EA_Die_Corner" : EA_Fail_corner,
                            "AI_Fail_EA_Die_Surface" : EA_Fail_die,
                            "AI_Fail_Image_Die_Corner" : crack_counter + chipout_counter + others_counter,
                            "AI_Fail_Image_Die_Surface" : incorrect_mag_counter,
                            "AI_Fail_Total" : total_ai_fail,
                            "Total_AOI_Die_Corner_Image" : ai_total_image-incorrect_mag_counter,
                            "AI_Pass" : data["AI_Pass"],
                            "AI_Reduction_Die_Corner" : ai_reduction_Sawpercent,
                            "AI_Reduction_All" : ai_reduction_percent,
                            "True_Fail" : len(data["OP_FailDetails"]),
                            "True_Fail_Crack" : True_Fail_Crack,
                            "True_Fail_Chipout" :True_Fail_Chipout,
                            "True_Fail_Die_Surface" : True_Fail_Die_Surface,
                            "True_Fail_Others" : True_Fail_Others,
                            "EA_True_Fail_Crack" : EA_OP_Crack ,
                            "EA_True_Fail_Chipout" : EA_OP_Chipout,
                            "EA_True_Fail_Die_Surface" : EA_OP_DieSurface,
                            "EA_True_Fail_Others": EA_OP_Others,
                            "EA_True_Fail_Crack_Chipout" : EA_OP_Crack+EA_OP_Chipout,
                            "Device_ID" : data_device_ID,
                            "OP_EA_Die_Corner" : len(OP_ChipOut),
                            "OP_EA_Die_Surface" : len(OP_Metal_Scratch),
                            "OP_EA_Others": len(OP_Others),
                            "Die_Overkill": EA_Fail_corner + EA_Fail_die - data["Reject_Amount"]
                        }
                        list_data.append(data_dictionary)

    for dic in list_data:
        now1 = datetime.datetime.strptime(dic["Date"], '%Y-%m-%d %H:%M:%S')
        if yesterday1 <= now1 <= today1:
            list_data_t.append(dic)

    list_data_t = sorted(list_data_t, key=lambda x: x["Date"])
    for dic in list_data_t:
        list_data2.append(dic)

    if weekend == "Weekend":
        yesterday = yesterday + "~" + (date_today - datetime.timedelta(1)).strftime("%m%d")

    wb, ws1 = reset_ws()
    excel_row = 2

    if list_data2:
        print("Creating csv")
        keys = list_data2[0].keys()
        directory_name = directory.split('\\')[-1]
        directory_name = directory_name.split("-")[1:]
        directory_name = "".join(directory_name)
        csv_path = output_path + "\\" + yesterday + "_All_(Security C)" + ".csv"

        # 根據 output_type 輸出 CSV 或 Excel 或兩個都輸出
        if output_type == "csv" or output_type == "both":
            if weekend != "Weekend":
                with open(csv_path, 'w', newline='') as output_file:
                    dict_writer = csv.DictWriter(output_file, keys)
                    dict_writer.writeheader()
                    dict_writer.writerows(list_data2)

        if output_type == "excel" or output_type == "both":
            for list in list_data2:
                side1 = Side(color='000000', style='thin')
                cells = ws1['A' + str(excel_row):'AX' + str(excel_row)]
                for cell in cells:
                    for cel in cell:
                        cel.font = Font(name='新細明體', size=12)
                        cel.alignment = Alignment(vertical='center', horizontal='center')
                        cel.border = Border(left=side1, right=side1, top=side1, bottom=side1)

                # 設定數字格式
                number_formats = {
                    'A': numbers.FORMAT_DATE_DATETIME,
                    'B': 'yyyy/mm/dd',
                    'H': '0.00%',
                    'I': '0.00%',
                    'L': '0.00%',
                    'M': '0.00%',
                    'P': '0.00%',
                    'Q': '0.00%',
                    'R': '0.00%',
                    'S': '0.00%',
                    'T': '0.00%',
                    'AH': '0.00%',
                    'AI': '0.00%'
                }
                for column, format in number_formats.items():
                    ws1[column + str(excel_row)].number_format = format

                # 設定資料
                data_mapping = {
                    'A': 'Date',
                    'B': 'Date_1',
                    'C': 'Lot',
                    'D': 'AOI_ID',
                    'E': 'AOI_Scan_Amount',
                    'F': 'AOI_Pass_Amount',
                    'G': 'AOI_Reject_Amount',
                    'H': 'AOI_Yield',
                    'I': 'AOI_Yield_Die_Corner',
                    'J': 'AI_Pass_Amount',
                    'K': 'AI_Reject_Amount',
                    'L': 'AI_Yield',
                    'M': 'AI_Fail_Corner_Yield',
                    'N': 'Final_Pass_Amount',
                    'O': 'Final_Reject_Amount',
                    'P': 'Final_Yield',
                    'Q': 'AI_EA_Overkill_Die_Corner',
                    'R': 'AI_EA_Overkill_Die_Surface',
                    'S': 'AI_Image_Overkill_Die_Corner',
                    'T': 'AI_Image_Overkill_Die_Surface',
                    'U': 'EA_over_kill_Die_Corner',
                    'V': 'EA_over_kill_Die_Surface',
                    'W': 'Image_Overkill_Die_Corner',
                    'X': 'Image_Overkill_Die_Surface',
                    'Y': 'Total_Images',
                    'Z': 'Image_Overkill',
                    'AA': 'AI_Fail_EA_Die_Corner',
                    'AB': 'AI_Fail_EA_Die_Surface',
                    'AC': 'AI_Fail_Image_Die_Corner',
                    'AD': 'AI_Fail_Image_Die_Surface',
                    'AE': 'AI_Fail_Total',
                    'AF': 'Total_AOI_Die_Corner_Image',
                    'AG': 'AI_Pass',
                    'AH': 'AI_Reduction_Die_Corner',
                    'AI': 'AI_Reduction_All',
                    'AJ': 'True_Fail',
                    'AK': 'True_Fail_Crack',
                    'AL': 'True_Fail_Chipout',
                    'AM': 'True_Fail_Die_Surface',
                    'AN': 'True_Fail_Others',
                    'AO': 'EA_True_Fail_Crack',
                    'AP': 'EA_True_Fail_Chipout',
                    'AQ': 'EA_True_Fail_Die_Surface',
                    'AR': 'EA_True_Fail_Others',
                    'AS': 'EA_True_Fail_Crack_Chipout',
                    'AT': 'Device_ID',
                    'AU': 'OP_EA_Die_Corner',
                    'AV': 'OP_EA_Die_Surface',
                    'AW': 'OP_EA_Others',
                    'AX': 'Die_Overkill'
                }
                for column, key in data_mapping.items():
                    ws1[column + str(excel_row)] = list[key]

                excel_row += 1

            # 匯出Excel
            while True:
                try:
                    wb.save(output_path + "\\" + yesterday + "_All_(Security C).xlsx")
                    break
                except Exception as error:
                    print(error)
                    time.sleep(1)

    print(directories)

# 將 CSV 資料寫入 MySQL 資料庫
def CsvToMysql(csv_folder, target_folder, db_host, db_user, db_password, db_name, table_name):
    """
    將 CSV 檔案匯入 MySQL 資料庫。

    Args:
        csv_folder (str): CSV 檔案所在資料夾路徑。
        target_folder (str): 目標資料夾路徑。
        db_host (str): MySQL 資料庫主機地址。
        db_user (str): MySQL 資料庫使用者名稱。
        db_password (str): MySQL 資料庫密碼。
        db_name (str): MySQL 資料庫名稱。
        table_name (str): MySQL 資料表名稱。
    """

    # 建立 MySQL 連線
    mydb = mysql.connector.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        database=db_name
    )

    # 驗證連線
    if mydb.is_connected():
        print("Database connection successful")
    else:
        print("Database connection failed")

    # 建立 Cursor 物件
    mycursor = mydb.cursor()

    # 建立 INSERT 語法
    sql = "INSERT INTO {} (Date, Date_1, Lot, AOI_ID, AOI_Scan_Amount, AOI_Pass_Amount, AOI_Reject_Amount, AOI_Yield, AOI_Yield_Die_Corner, AI_Pass_Amount, AI_Reject_Amount, AI_Yield, AI_Fail_Corner_Yield, Final_Pass_Amount, Final_Reject_Amount, Final_Yield, AI_EA_Overkill_Die_Corner, AI_EA_Overkill_Die_Surface, AI_Image_Overkill_Die_Corner, AI_Image_Overkill_Die_Surface, EA_over_kill_Die_Corner, EA_over_kill_Die_Surface, Image_Overkill_Die_Corner, Image_Overkill_Die_Surface, Total_Images, Image_Overkill, AI_Fail_EA_Die_Corner, AI_Fail_EA_Die_Surface, AI_Fail_Image_Die_Corner, AI_Fail_Image_Die_Surface, AI_Fail_Total, Total_AOI_Die_Corner_Image, AI_Pass, AI_Reduction_Die_Corner, AI_Reduction_All, True_Fail, True_Fail_Crack, True_Fail_Chipout, True_Fail_Die_Surface, True_Fail_Others, EA_True_Fail_Crack, EA_True_Fail_Chipout, EA_True_Fail_Die_Surface, EA_True_Fail_Others, `EA_True_Fail_Crack_Chipout`, Device_ID, OP_EA_Die_Corner, OP_EA_Die_Surface, OP_EA_Others, Die_Overkill) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)".format(table_name)

    # 初始化總插入筆數
    total_rows_inserted = 0

    # 遍歷資料夾中的檔案
    for filename in os.listdir(csv_folder):
        # 使用正則表達式檢查檔案名稱是否符合命名規則
        if re.match(r'^\d{2}\d{2}_All_\(Security C\).csv$', filename):
            # 構建完整檔案路徑
            csv_file = os.path.join(csv_folder, filename)

            # 讀取 CSV 檔案
            df = pd.read_csv(csv_file)

            # 將 DataFrame 轉換成資料列表
            data = df.values.tolist()

            # 執行 INSERT 語法
            try:
                mycursor.executemany(sql, data)
                mydb.commit()

                # 取得插入的資料筆數
                rows_inserted = mycursor.rowcount
                total_rows_inserted += rows_inserted
                print(f"File: {filename} IN {rows_inserted}!")

                # 取得檔案月份
                month = filename[:2]

                # 取得當年的年份
                current_year = datetime.datetime.now().year

                # 構建目標資料夾路徑
                month_folder = os.path.join(target_folder, str(current_year), month)

                # 檢查資料夾是否存在，不存在則建立資料夾
                if not os.path.exists(month_folder):
                    os.makedirs(month_folder)

                # 移動檔案
                shutil.move(csv_file, month_folder)
            except mysql.connector.Error as error:
                print(f"File: {filename} ERROR: {error}")

    # 輸出總共插入的資料筆數
    print(f"\nTotal rows inserted: {total_rows_inserted}")

    # 關閉 Cursor 和連線
    mycursor.close()
    mydb.close()

# 將 Excel Results 中的歷史 CSV 轉換為可入資料庫的型態
def TransformHistoricalData(old_file_path, new_file_path):
    """
    處理 CSV 檔案，更換欄位名稱並新增 Die_Overkill 欄位。

    Args:
        old_file_path (str): 舊 CSV 檔案的路徑。
        new_file_path (str): 新 CSV 檔案的路徑。
    """

    # 偵測檔案編碼
    with open(old_file_path, 'rb') as f:
        encoding = chardet.detect(f.read())['encoding']

    # 讀取舊檔案，指定編碼
    df = pd.read_csv(old_file_path, encoding=encoding)

    # 欄位名稱對照表
    column_mapping = {
        'Date': 'Date',
        'Date-1': 'Date_1',
        'Lot': 'Lot',
        'ID': 'AOI_ID',
        'AOI scan amount': 'AOI_Scan_Amount',
        'AOI pass amount': 'AOI_Pass_Amount',
        'AOI reject amount': 'AOI_Reject_Amount',
        'AOI yield(%)': 'AOI_Yield',
        'AOI yield(%) Die Corner%': 'AOI_Yield_Die_Corner',
        'AI pass amount': 'AI_Pass_Amount',
        'AI reject amount': 'AI_Reject_Amount',
        'AI yield(%)': 'AI_Yield',
        'AI Fail corner yield(%)': 'AI_Fail_Corner_Yield',
        'Final pass amount': 'Final_Pass_Amount',
        'Final reject amount': 'Final_Reject_Amount',
        'Final yield(%)': 'Final_Yield',
        'AI EA Overkill Die Corner%': 'AI_EA_Overkill_Die_Corner',
        'AI EA Overkill Die Surface%': 'AI_EA_Overkill_Die_Surface',
        'AI Image Overkill Die Corner%': 'AI_Image_Overkill_Die_Corner',
        'AI Image Overkill Die Surface%': 'AI_Image_Overkill_Die_Surface',
        'EA over-kill Die Corner': 'EA_over_kill_Die_Corner',
        'EA over-kill Die Surface': 'EA_over_kill_Die_Surface',
        'Image over-kill Die Corner': 'Image_Overkill_Die_Corner',
        'Image over-kill Die Surface': 'Image_Overkill_Die_Surface',
        'Total Images': 'Total_Images',
        'Image over-kill': 'Image_Overkill',
        'AI Fail EA Die Corner': 'AI_Fail_EA_Die_Corner',
        'AI Fail EA Die Surface': 'AI_Fail_EA_Die_Surface',
        'AI Fail Image Die Corner': 'AI_Fail_Image_Die_Corner',
        'AI Fail Image Die Surface': 'AI_Fail_Image_Die_Surface',
        'AI Fail Total': 'AI_Fail_Total',
        'Total AOI Die Corner image': 'Total_AOI_Die_Corner_Image',
        'AI Pass': 'AI_Pass',
        'AI reduction Die Corner(%)': 'AI_Reduction_Die_Corner',
        'AI reduction All(%)': 'AI_Reduction_All',
        'True fail': 'True_Fail',
        'True fail Crack': 'True_Fail_Crack',
        'True fail Chipout': 'True_Fail_Chipout',
        'True fail Die Surface': 'True_Fail_Die_Surface',
        'True fail Others': 'True_Fail_Others',
        'EA True fail Crack': 'EA_True_Fail_Crack',
        'EA True fail Chipout': 'EA_True_Fail_Chipout',
        'EA True fail Die Surface': 'EA_True_Fail_Die_Surface',
        'EA True fail Others': 'EA_True_Fail_Others',
        'EA True fail Crack+Chipout': 'EA_True_Fail_Crack_Chipout',
        'Device ID': 'Device_ID',
        'OP EA Die Corner': 'OP_EA_Die_Corner',
        'OP EA Die Surface': 'OP_EA_Die_Surface',
        'OP EA Others': 'OP_EA_Others'
    }

    # 更換欄位名稱
    df = df.rename(columns=column_mapping)

    # 新增 Die_Overkill 欄位
    df['Die_Overkill'] = df['AI_Reject_Amount'] - df['Final_Reject_Amount']

    # 儲存新檔案
    df.to_csv(new_file_path, index=False, encoding=encoding)

    print(f"CSV file updated and saved to {new_file_path}")

# ----------------------------------- 參數設定 -----------------------------------

# 要切換的 AI_Result\Excel Results 歷史檔案日期
set_date = "0630"

# AOI驗證測試\AI_Result\Excel Results 檔案路徑
old_file_path = r"\\KHFS2\WBG PE Stage$\AOI 判圖\AOI驗證測試\AI_Result\Excel Results\today\old csv\All_2OAOI_{}_(Security C).csv".format(set_date)

# 切換正式或測試環境的資料讀取路徑
env = "prod"  # 環境變數

if env == "dev":
    settings_path = r"\\khwbpeaiaoi01\2451AOI$\WaferMapTemp\AI_Result - Copy\settings.json"
    main_path = r"\\khwbpeaiaoi01\2451AOI$\WaferMapTemp\AI_Result - Copy"
    csv_folder = "D:\ASEKH\K18330\DataProcess"
    target_folder = "D:\ASEKH\K18330\DataProcess\All Data"
    new_file_path = r"D:\ASEKH\K18330\DataProcess\{}_All_(Security C).csv".format(set_date)
    db_host = '127.0.0.1'
elif env == "prod":
    settings_path = r"\\khwbpeaiaoi01\2451AOI$\WaferMapTemp\AI_Result\settings\settings.json"
    main_path = r"\\khwbpeaiaoi01\2451AOI$\WaferMapTemp\AI_Result"
    csv_folder = r"\\khwbpeaiaoi01\D$\khwbpeaiaoi_Shares$\K18330\DataBase"
    target_folder = r"\\khwbpeaiaoi01\D$\khwbpeaiaoi_Shares$\K18330\DataBase\All Data"
    new_file_path = r"\\khwbpeaiaoi01\D$\khwbpeaiaoi_Shares$\K18330\DataBase\{}_All_(Security C).csv".format(set_date)
    db_host = '10.11.33.122'
else:
    print("請設定正確的環境變數：dev 或 prod")
    exit()

# MySQL 連線資訊
db_user = 'root'
db_password = ''
db_name = 'wb'

# 資料表名稱
table_name = 'all_2oaoi'

# 初始化
wb, ws1 = reset_ws()
database = read_Jsonfile(settings_path)
now = datetime.datetime.now()

# ----------------------------------- 主程式 -----------------------------------

    # 三種使用模式：
    #     寫入當天資料：自動抓取當日資料寫入資料庫。
    #     寫入過去資料：設置日期(0701 - 0702)，抓取 0701、0702 資料寫入資料庫。
    #     處理歷史 CSV 檔案：使用 AOI驗證測試\AI_Result\Excel Results 中的歷史資料轉檔入資料庫。

# 寫入當日資料
JsonToExcel(database, main_path, csv_folder,(now + datetime.timedelta(-1)).strftime('%m%d'), now.strftime('%m%d'), wb, ws1, output_type="csv")
CsvToMysql(csv_folder, target_folder, db_host, db_user, db_password, db_name, table_name)

# 寫入過去資料 start_day ~ end_day
# start_day = "0615"
# end_day = "0616"
# for date in range(int(start_day), int(end_day) + 1):
#     start_date = str(date).zfill(4)
#     end_date = str(date + 1).zfill(4)
#     print(start_date)
#     JsonToExcel(database, main_path, csv_folder, start_date, end_date, wb, ws1, output_type="csv")
#     CsvToMysql(csv_folder, target_folder, db_host, db_user, db_password, db_name, table_name)

# 處理歷史 CSV 檔案
# TransformHistoricalData(old_file_path, new_file_path)
# CsvToMysql(csv_folder, target_folder, db_host, db_user, db_password, db_name, table_name)