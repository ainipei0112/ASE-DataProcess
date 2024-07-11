import os
import csv
import json
import time
import numpy as np
import shutil
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill,numbers
import pandas as pd
import mysql.connector
import re

# ----------------------------------- 函數定義 -----------------------------------

# 設定 Excel 標題格式
def resetws():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = '工作表1'
    # 標題欄位
    column_titles = ['Date', 'Date_1', 'Lot', 'AOI_ID', 'AOI_Scan_Amount', 'AOI_Pass_Amount', 'AOI_Reject_Amount', 
                    'AOI_Yield', 'AOI_Yield_Die_Corner', 'AI_Pass_Amount', 'AI_Reject_Amount', 'AI_Yield', 
                    'AI_Fail_Corner_Yield', 'Final_Pass_Amount', 'Final_Reject_Amount', 'Final_Yield', 
                    'AI_EA_Overkill_Die_Corner', 'AI_EA_Overkill_Die_Surface', 'AI_Image_Overkill_Die_Corner', 
                    'AI_Image_Overkill_Die_Surface', 'EA_over_kill_Die_Corner', 'EA_over_kill_Die_Surface', 
                    'Image_Overkill_Die_Corner', 'Image_Overkill_Die_Surface', 'Total_Images', 'Image_Overkill', 
                    'AI_Fail_EA_Die_Corner', 'AI_Fail_EA_Die_Surface', 'AI_Fail_Image_Die_Corner', 
                    'AI_Fail_Image_Die_Surface', 'AI_Fail_Total', 'Total_AOI_Die_Corner_Image', 'AI_Pass', 
                    'AI_Reduction_Die_Corner', 'AI_Reduction_All', 'True_Fail', 'True_Fail_Crack', 'True_Fail_Chipout', 
                    'True_Fail_Die_Surface', 'True_Fail_Others', 'EA_True_Fail_Crack', 'EA_True_Fail_Chipout', 
                    'EA_True_Fail_Die_Surface', 'EA_True_Fail_Others', 'EA_True_Fail_Crack_Chipout', 'Device_ID', 
                    'OP_EA_Die_Corner', 'OP_EA_Die_Surface', 'OP_EA_Others', 'Die_Overkill']
    for i, title in enumerate(column_titles, start=1):
        ws1.cell(row=1, column=i, value=title)
        ws1.column_dimensions[get_column_letter(i)].width = 22
        fill_color = '95B3D7' if i < 21 else ('FDE9D9' if i < 27 else 'B7DEE8')
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        ws1.cell(row=1, column=i).fill = fill
        side1 = Side(color='000000', style='thin')
        ws1.cell(row=1, column=i).font = Font(name='Microsoft YaHei', size=12)
        ws1.cell(row=1, column=i).alignment = Alignment(vertical='center', horizontal='center')
        ws1.cell(row=1, column=i).border = Border(left=side1, right=side1, top=side1, bottom=side1)
    ws1.row_dimensions[1].height = 49.5
    return wb, ws1

# 讀取資料庫設定
def read_database_settings(settings_path):
    try:
        with open(settings_path, "r", encoding='utf-8') as r_file:
            databases = json.load(r_file)
        database = [data for data in databases["folder_details"]]
        return database
    except FileNotFoundError:
        print(f"Error: Settings file '{settings_path}' not found.")
        exit()
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON format in '{settings_path}'.")
        exit()
    except Exception as e:
        print(f"Error reading database settings: {e}")
        exit()

# 處理 JSON 資料並寫入 Excel
def process_json_data(directory, database, yesterday1, today1, wb, ws1):
    list_data = []
    list_data2 = []
    list_data_t = []
    excel_row = 2

    # 遍歷資料夾中的 JSON 檔案
    for filename in os.listdir(directory):
        if filename.endswith(".json"):
            try:
                with open(os.path.join(directory, filename), "r", encoding='utf-8') as json_file:
                    data = json.load(json_file)
            except Exception as e:
                print(f"Error reading JSON file '{filename}': {e}")
                continue

            # 檢查 JSON 檔案是否符合條件
            if "OP_Checked" in data and data["AOI_Scan_Amount"] != 0:
                duplicate_Checker = {}
                xdie = []
                OG_loc = os.path.join(directory, filename).split("\\")
                date_obj = datetime.datetime.strptime(OG_loc[6], "%Y-%m-%d")
                date_str = date_obj.strftime("%m%d%Y")
                OG_Loc = '\\\\' + OG_loc[2] + '\\' + OG_loc[3] + '\\' + OG_loc[4] + '\\Image' + '\\' + date_str + '\\' + OG_loc[7] + '\\' + OG_loc[8].replace('.json', '')

                # 檢查資料夾是否存在
                if not os.path.exists(OG_Loc):
                    print('DIE', OG_Loc)
                    continue
                else:
                    print('OK', OG_Loc)

                data_filename = filename.split(".")[:-1]
                data_filename = ".".join(data_filename)

                # 根據檔案名稱查找資料庫中的相關資訊
                for database_data in database:
                    if database_data["filename"] == data_filename:
                        data_date = database_data["Date"]
                        data_device_ID = database_data["Device_ID"]
                        break

                if int(data["AOI_Scan_Amount"]) == 0:
                    continue

                # 根據時間區間篩選資料
                date_time = datetime.datetime.strptime(data_date, "%Y-%m-%d %H:%M:%S").time()
                compare_time = datetime.time(7, 30)

                if date_time >= compare_time:
                    date_day = datetime.datetime.strptime(data_date, "%Y-%m-%d %H:%M:%S").date()
                else:
                    date_day = (datetime.datetime.strptime(data_date, "%Y-%m-%d %H:%M:%S") - datetime.timedelta(1)).date()

                # 建立資料字典
                data_dictionary = {
                    "Date": data_date,
                    "Date_1": date_day,
                    "Lot": directory.split("\\")[-1],
                    "AOI_ID": filename.split(".")[-2],
                    "AOI_Scan_Amount": data["AOI_Scan_Amount"]
                }
                list_data.append(data_dictionary)

    # 根據時間區間篩選資料
    for dic in list_data:
        now1 = datetime.datetime.strptime(dic["Date"], '%Y-%m-%d %H:%M:%S')
        if yesterday1 <= now1 <= today1:
            list_data_t.append(dic)

    # 排序資料
    list_data_t = sorted(list_data_t, key=lambda x: x["Date"])
    for dic in list_data_t:
        list_data2.append(dic)

    # 將資料寫入 Excel
    if list_data2:
        print("Creating csv")
        keys = list_data2[0].keys()
        directory_name = directory.split('\\')[-1]
        directory_name = directory_name.split("-")[1:]
        directory_name = "".join(directory_name)
        csv_path = r'D:\ASEKH\K18330\DataProcess' + "\\" + yesterday1.strftime('%m%d') + "_All_(Security C)" + ".csv"
        # csv_path = r'\\10.11.33.122\D$\khwbpeaiaoi_Shares$\K18330\DataBase' + "\\" + yesterday1.strftime('%m%d') + "_All_(Security C)" + ".csv"
        with open(csv_path, 'w', newline='') as output_file:
            dict_writer = csv.DictWriter(output_file, keys)
            dict_writer.writeheader()
            dict_writer.writerows(list_data2)

    return wb, ws1

# 將 Excel 檔案資料寫入資料庫
def write_excel_to_database(excel_folder, db_host, db_user, db_password, db_name, table_name):
    mydb = mysql.connector.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        database=db_name
    )

    mycursor = mydb.cursor()

    sql = "INSERT INTO {} (Date, Date_1, Lot, AOI_ID, AOI_Scan_Amount, AOI_Pass_Amount, AOI_Reject_Amount, AOI_Yield, AOI_Yield_Die_Corner, AI_Pass_Amount, AI_Reject_Amount, AI_Yield, AI_Fail_Corner_Yield, Final_Pass_Amount, Final_Reject_Amount, Final_Yield, AI_EA_Overkill_Die_Corner, AI_EA_Overkill_Die_Surface, AI_Image_Overkill_Die_Corner, AI_Image_Overkill_Die_Surface, EA_over_kill_Die_Corner, EA_over_kill_Die_Surface, Image_Overkill_Die_Corner, Image_Overkill_Die_Surface, Total_Images, Image_Overkill, AI_Fail_EA_Die_Corner, AI_Fail_EA_Die_Surface, AI_Fail_Image_Die_Corner, AI_Fail_Image_Die_Surface, AI_Fail_Total, Total_AOI_Die_Corner_Image, AI_Pass, AI_Reduction_Die_Corner, AI_Reduction_All, True_Fail, True_Fail_Crack, True_Fail_Chipout, True_Fail_Die_Surface, True_Fail_Others, EA_True_Fail_Crack, EA_True_Fail_Chipout, EA_True_Fail_Die_Surface, EA_True_Fail_Others, `EA_True_Fail_Crack_Chipout`, Device_ID, OP_EA_Die_Corner, OP_EA_Die_Surface, OP_EA_Others, Die_Overkill) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)".format(table_name)

    total_rows_inserted = 0

    # 遍歷資料夾中的檔案
    for filename in os.listdir(excel_folder):
        if re.match(r'^\d{2}\d{2}_All_\(Security C\).xlsx$', filename):
            excel_file = os.path.join(excel_folder, filename)

            try:
                df = pd.read_excel(excel_file)
                data = df.values.tolist()
                mycursor.executemany(sql, data)
                mydb.commit()
                rows_inserted = mycursor.rowcount
                total_rows_inserted += rows_inserted
                print(f"File: {filename} IN {rows_inserted}!")
            except mysql.connector.Error as error:
                print(f"File: {filename} ERROR: {error}")
            except Exception as e:
                print(f"Error reading Excel file '{filename}': {e}")
                continue

    print(f"\nTotal rows inserted: {total_rows_inserted}")
    mycursor.close()
    mydb.close()

# ----------------------------------- 主程式 -----------------------------------

# 設定資料庫連線資訊
db_host = '127.0.0.1'
db_user = 'root'
db_password = ''
db_name = 'wb'
table_name = 'all_2oaoi'

# 設定資料夾路徑
settings_path = r"\\khwbpeaiaoi01\2451AOI$\WaferMapTemp\AI_Result - Copy\settings.json"
main_path = r"\\khwbpeaiaoi01\2451AOI$\WaferMapTemp\AI_Result - Copy"
excel_folder = "C:\\Users\K18330\Desktop\\DataProcess"

# 讀取資料庫設定
database = read_database_settings(settings_path)

# 獲取當前時間
now = datetime.datetime.now()
yesterday = now + datetime.timedelta(-1)

# 處理 JSON 資料並寫入 Excel
wb, ws1 = resetws()
wb, ws1 = process_json_data(main_path, database, yesterday, now, wb, ws1)

# 將 Excel 檔案資料寫入資料庫
write_excel_to_database(excel_folder, db_host, db_user, db_password, db_name, table_name)

print("處理完成！")